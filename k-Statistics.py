from timeit import default_timer as timer
from openpyxl import Workbook
import pandas as pd
import numpy as np

def regular_power_sum(num, arr):
    total = 0
    for i in range (0, len(arr)):
        total = total + arr[i]**num
    return total

def sumOfProduct(arr, n, k):
     
    # Initialising all the values to 0
    dp = [ [ 0 for x in range(n + 1)] for y in range(n + 1)]
     
    # To store the answer for
    # current value of k
    cur_sum = 0
     
    # For k = 1, the answer will simply
    # be the sum of all the elements
    for i in range(1, n + 1):
        dp[1][i] = arr[i - 1]
        cur_sum += arr[i - 1]
  
    # Filling the table in bottom up manner
    for i in range(2 , k + 1):
  
        # To store the elements of the current
        # row so that we will be able to use this sum
        # for subsequent values of k
        temp_sum = 0
  
        for j in range( 1,  n + 1):
  
            # We will subtract previously computed value
            # so as to get the sum of elements from j + 1
            # to n in the (i - 1)th row
            cur_sum -= dp[i - 1][j]
  
            dp[i][j] = arr[j - 1] * cur_sum
            temp_sum += dp[i][j]
        cur_sum = temp_sum
    return cur_sum


read = pd.read_csv('DataMM.csv')
data = read.to_numpy()
values = data[:,0].tolist()
print(sum(values) / len(values))
size = len(values)

wb = Workbook()
ws1 = wb.active

# for i in range (4, size):
#     for j in range (0, 20):
#         arr = [val * 10**j for val in values[0:i+1]]
#         start_time = timer()
#         power_sum = []
#         for k in range (0, 4):
#             power_sum.append(regular_power_sum(k+1, arr))
#         end_time = timer()
#         ws1.cell(row=i+1-4, column=j+1).value = 10**6 * (end_time - start_time)

for i in range (1, 11):
    arr = values[0:5*i]
    start_time = timer()
    power_sum = []
    for k in range (0, 5):
        power_sum.append(regular_power_sum(k+1, arr))
    end_time = timer()
    ws1.cell(row=1, column=i).value = 10**6 * (end_time - start_time)

for i in range (1, 11):
    arr = values[0:5*i]
    el_sym_sum = []
    for k in range (0, 5*i):
        el_sym_sum.append(sumOfProduct(arr, 5*i, k+1))
    pow_sum = []
    start_time = timer()
    for k in range (0, 5):
        term = (-1)**k * (k+1) * el_sym_sum[k]
        for p in range (0, k):
            term = term + (-1)**p * el_sym_sum[p] * pow_sum[k-p-1]
        pow_sum.append(term)
    end_time = timer()
    ws1.cell(row=2, column=i).value = 10**6 * (end_time - start_time)


# ws2 = wb.create_sheet()

# for i in range (4, size):
#     for j in range (0, 20):
#         arr = [val * 10**j for val in values[0:i+1]]
#         el_sym_sum = []
#         for k in range (0, i+1):
#             el_sym_sum.append(sumOfProduct(arr, i+1, k+1))
#         pow_sum = []
#         start_time = timer()
#         for k in range (0, 4):
#             term = (-1)**k * (k+1) * el_sym_sum[k]
#             for p in range (0, k):
#                 term = term + (-1)**p * el_sym_sum[p] * pow_sum[k-p-1]
#             pow_sum.append(term)
#         end_time = timer()
#         ws2.cell(row=i+1-4, column=j+1).value = 10**6 * (end_time - start_time)




wb.save("regular.xlsx")

el_sym_sum = []
for i in range (0, size):
    el_sym_sum.append(sumOfProduct(values, size, i+1))

pow_sum = []
for i in range (0, size):
    term = (-1)**i * (i+1) * el_sym_sum[i]
    for j in range (0, i):
        term = term + (-1)**j * el_sym_sum[j] * pow_sum[i-j-1]
    pow_sum.append(term)


#start_time = time.time()
power_sum = []
for k in range (0, 70):
    power_sum.append(regular_power_sum(k+1, arr))
#end_time = time.time()

values = [69, 73, 84, 89, 97]

# print(sumOfProduct(values, 5, 5))
print (regular_power_sum(4, values))